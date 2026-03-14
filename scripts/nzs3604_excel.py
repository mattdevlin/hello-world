#!/usr/bin/env python3
"""
NZS 3604 Excel Report Generator

Generates an Excel workbook with 3 sheets:
1. "NZS 3604 Timber Schedule" — all member sizes per level/element
2. "NZS 3604 Fixing Schedule" — nailing/bolting per connection type
3. "Stickframe vs SIP Comparison" — side-by-side comparison table

Dependencies: openpyxl

Usage:
    python nzs3604_excel.py nzs3604_design.json [comparison.json] [output.xlsx]
"""

import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Calibri", bold=True, size=11)
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_FONT_WHITE = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
SECTION_FONT = Font(name="Calibri", bold=True, size=10, color="003366")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_header(ws, row, cols, headers):
    """Apply header styling to a row."""
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    return row + 1


def _set_col_widths(ws, widths):
    """Set column widths."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 1: Timber Schedule
# ---------------------------------------------------------------------------


def write_timber_schedule(ws, design):
    """Write the NZS 3604 timber schedule sheet."""
    ws.title = "NZS 3604 Timber Schedule"
    _set_col_widths(ws, [15, 20, 15, 15, 15, 20, 25])

    row = 1
    # Title
    ws.cell(row=row, column=1, value="NZS 3604:2011 Timber Framing Schedule — SG8")
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14)
    row += 2

    # Site info
    site = design.get("site", {})
    ws.cell(row=row, column=1, value="Wind Zone:")
    ws.cell(row=row, column=2, value=site.get("wind_zone", ""))
    row += 1
    ws.cell(row=row, column=1, value="EQ Zone:")
    ws.cell(row=row, column=2, value=str(site.get("eq_zone", "")))
    row += 1
    ws.cell(row=row, column=1, value="Soil Type:")
    ws.cell(row=row, column=2, value=site.get("soil_type", ""))
    row += 2

    headers = ["Level", "Element", "Size", "Spacing", "Max Span", "Fixing", "Reference"]
    row = _apply_header(ws, row, 7, headers)

    for level in design.get("levels", []):
        level_name = level.get("level_name", "")

        # Walls
        for wall in level.get("walls", []):
            wall_name = wall.get("wall_id", wall.get("wall_name", "Wall"))
            stud = wall.get("stud", {})
            ws.cell(row=row, column=1, value=level_name)
            ws.cell(row=row, column=2, value=f"Wall {wall_name} — Studs")
            ws.cell(row=row, column=3, value=stud.get("size", ""))
            ws.cell(row=row, column=4, value=f"{stud.get('spacing', '')}mm")
            ws.cell(row=row, column=6, value=stud.get("ref", ""))
            for c in range(1, 8):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

            tp = wall.get("top_plate", {})
            ws.cell(row=row, column=2, value=f"Wall {wall_name} — Top Plate")
            ws.cell(row=row, column=3, value=tp.get("size", ""))
            ws.cell(row=row, column=5, value=f"{tp.get('max_span_m', '')}m" if tp.get("max_span_m") else "")
            ws.cell(row=row, column=7, value=tp.get("ref", ""))
            for c in range(1, 8):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

            bp = wall.get("bottom_plate", {})
            ws.cell(row=row, column=2, value=f"Wall {wall_name} — Bottom Plate")
            ws.cell(row=row, column=3, value=bp.get("size", ""))
            ws.cell(row=row, column=5, value=f"{bp.get('max_span_m', '')}m" if bp.get("max_span_m") else "")
            ws.cell(row=row, column=7, value=bp.get("ref", ""))
            for c in range(1, 8):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

        # Floor
        floor = level.get("floor", {})
        if floor.get("type") == "suspended_timber":
            for zone in floor.get("zones", []):
                joist = zone.get("joist", {})
                ws.cell(row=row, column=1, value=level_name)
                ws.cell(row=row, column=2, value=f"Floor — {zone.get('zone', '')} Joists")
                ws.cell(row=row, column=3, value=joist.get("size", ""))
                ws.cell(row=row, column=4, value=f"{joist.get('spacing', '')}mm")
                ws.cell(row=row, column=5, value=f"{joist.get('max_span_m', '')}m" if joist.get("max_span_m") else "")
                ws.cell(row=row, column=7, value=joist.get("ref", ""))
                for c in range(1, 8):
                    ws.cell(row=row, column=c).border = THIN_BORDER
                row += 1

                bearer = zone.get("bearer", {})
                ws.cell(row=row, column=2, value=f"Floor — {zone.get('zone', '')} Bearers")
                ws.cell(row=row, column=3, value=bearer.get("size", ""))
                ws.cell(row=row, column=7, value=bearer.get("ref", ""))
                for c in range(1, 8):
                    ws.cell(row=row, column=c).border = THIN_BORDER
                row += 1

        # Roof
        roof = level.get("roof")
        if roof:
            rafter = roof.get("rafter", {})
            ws.cell(row=row, column=1, value=level_name)
            ws.cell(row=row, column=2, value="Roof — Rafters")
            ws.cell(row=row, column=3, value=rafter.get("size", ""))
            ws.cell(row=row, column=4, value=f"{rafter.get('spacing', '')}mm")
            ws.cell(row=row, column=5, value=f"{rafter.get('max_span_m', '')}m" if rafter.get("max_span_m") else "")
            ws.cell(row=row, column=6, value=rafter.get("fixing", ""))
            ws.cell(row=row, column=7, value=rafter.get("ref", ""))
            for c in range(1, 8):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

            cj = roof.get("ceiling_joist", {})
            ws.cell(row=row, column=2, value="Roof — Ceiling Joists")
            ws.cell(row=row, column=3, value=cj.get("size", ""))
            ws.cell(row=row, column=4, value=f"{cj.get('spacing', '')}mm")
            ws.cell(row=row, column=5, value=f"{cj.get('max_span_m', '')}m" if cj.get("max_span_m") else "")
            ws.cell(row=row, column=7, value=cj.get("ref", ""))
            for c in range(1, 8):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1

            ridge = roof.get("ridge_beam", {})
            if ridge.get("size") and ridge["size"] != "ridge board only":
                ws.cell(row=row, column=2, value="Roof — Ridge Beam")
                ws.cell(row=row, column=3, value=ridge.get("size", ""))
                ws.cell(row=row, column=5, value=f"{ridge.get('max_span_m', '')}m" if ridge.get("max_span_m") else "")
                ws.cell(row=row, column=6, value=ridge.get("fixing", ""))
                ws.cell(row=row, column=7, value=ridge.get("ref", ""))
                for c in range(1, 8):
                    ws.cell(row=row, column=c).border = THIN_BORDER
                row += 1

    return ws


# ---------------------------------------------------------------------------
# Sheet 2: Fixing Schedule
# ---------------------------------------------------------------------------


def write_fixing_schedule(ws, design):
    """Write the NZS 3604 fixing schedule sheet."""
    ws.title = "NZS 3604 Fixing Schedule"
    _set_col_widths(ws, [20, 30, 15, 10, 20, 25])

    row = 1
    ws.cell(row=row, column=1, value="NZS 3604:2011 Nailing & Fixing Schedule — SG8")
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14)
    row += 2

    headers = ["Section", "Joint", "Nail Size", "Qty", "Location", "Reference"]
    row = _apply_header(ws, row, 6, headers)

    connections = design.get("connections", {})

    for section_key, section_name in [
        ("subfloor", "S6 Subfloor"),
        ("floor", "S7 Floor"),
        ("walls", "S8 Walls"),
        ("roof", "S10 Roof"),
        ("ceiling", "S13 Ceiling"),
    ]:
        section = connections.get(section_key, {})
        joints = section.get("joints", section)
        ref = section.get("ref", "")

        for joint_name, joint_data in joints.items():
            if joint_name.startswith("_") or joint_name == "ref":
                continue

            # Handle cross-references
            if isinstance(joint_data, dict) and "_see" in joint_data:
                ws.cell(row=row, column=1, value=section_name)
                ws.cell(row=row, column=2, value=joint_name.replace("_", " ").title())
                ws.cell(row=row, column=3, value=f"See {joint_data['_see']}")
                ws.cell(row=row, column=6, value=ref)
                for c in range(1, 7):
                    ws.cell(row=row, column=c).border = THIN_BORDER
                row += 1
                continue

            # Extract hand nailing (preferred for schedule)
            hand = joint_data.get("hand") if isinstance(joint_data, dict) else None
            if hand is None:
                continue

            # Hand can be a single dict or array of alternatives
            if isinstance(hand, list):
                hand = hand[0]  # use first option

            if isinstance(hand, dict):
                ws.cell(row=row, column=1, value=section_name)
                ws.cell(row=row, column=2, value=joint_name.replace("_", " ").title())
                ws.cell(row=row, column=3, value=hand.get("nail", ""))
                ws.cell(row=row, column=4, value=str(hand.get("qty", "")))
                ws.cell(row=row, column=5, value=hand.get("location", ""))
                ws.cell(row=row, column=6, value=ref)
                for c in range(1, 7):
                    ws.cell(row=row, column=c).border = THIN_BORDER
                row += 1

    # Fixing types reference
    row += 1
    ws.cell(row=row, column=1, value="Fixing Types Reference (Table 10.15)")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    fixing_ref = connections.get("fixing_types_reference", {}).get("types", {})
    for fix_type, info in fixing_ref.items():
        ws.cell(row=row, column=1, value=f"Type {fix_type}")
        ws.cell(row=row, column=2, value=info.get("description", ""))
        ws.cell(row=row, column=3, value=f"{info.get('capacity_kN', '')} kN")
        for c in range(1, 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    return ws


# ---------------------------------------------------------------------------
# Sheet 3: Comparison
# ---------------------------------------------------------------------------


def write_comparison(ws, comparison):
    """Write the stickframe vs SIP comparison sheet."""
    ws.title = "Stickframe vs SIP"
    _set_col_widths(ws, [20, 30, 30, 25])

    row = 1
    ws.cell(row=row, column=1, value="NZS 3604 Stickframe vs DEVPRO SIP Comparison")
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14)
    row += 2

    headers = ["Element", "NZS 3604 Stickframe", "DEVPRO SIP", "Advantage"]
    row = _apply_header(ws, row, 4, headers)

    # Walls
    for wall_comp in comparison.get("walls", []):
        sf = wall_comp.get("stickframe", {})
        sip = wall_comp.get("devpro_sip", {})
        ws.cell(row=row, column=1, value=wall_comp.get("element", ""))
        ws.cell(row=row, column=2, value=f"{sf.get('component_count', '?')} pieces ({sf.get('stud_size', '')} studs)")
        ws.cell(row=row, column=3, value=f"{sip.get('panel_count', '?')} panels" if sip.get("panel_count") else "N/A")
        ws.cell(row=row, column=4, value=wall_comp.get("winner", ""))
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Floor
    floor = comparison.get("floor", {})
    if floor:
        ws.cell(row=row, column=1, value=floor.get("element", "Floor"))
        sf = floor.get("stickframe", {})
        ws.cell(row=row, column=2, value=sf.get("type", ""))
        ws.cell(row=row, column=3, value=floor.get("devpro_sip", {}).get("type", ""))
        ws.cell(row=row, column=4, value=floor.get("winner", ""))
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Roof
    roof = comparison.get("roof", {})
    if roof:
        ws.cell(row=row, column=1, value=roof.get("element", "Roof"))
        sf = roof.get("stickframe", {})
        ws.cell(row=row, column=2, value=f"{sf.get('rafter_size', '')} rafters @ {sf.get('rafter_spacing_mm', '')}mm")
        ws.cell(row=row, column=3, value=roof.get("devpro_sip", {}).get("type", ""))
        ws.cell(row=row, column=4, value=roof.get("winner", ""))
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Summary
    row += 1
    summary = comparison.get("summary", {})
    ws.cell(row=row, column=1, value="SUMMARY")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1

    sf_total = summary.get("stickframe", {}).get("total_wall_components", 0)
    sip_total = summary.get("devpro_sip", {}).get("total_wall_panels", 0)
    ws.cell(row=row, column=1, value="Total wall components:")
    ws.cell(row=row, column=2, value=str(sf_total))
    ws.cell(row=row, column=3, value=str(sip_total))
    row += 1

    reduction = summary.get("component_reduction_pct")
    if reduction:
        ws.cell(row=row, column=1, value="Component reduction:")
        ws.cell(row=row, column=2, value=f"{reduction}%")
        row += 1

    trade_red = summary.get("trade_reduction", 0)
    ws.cell(row=row, column=1, value="Trade reduction:")
    ws.cell(row=row, column=2, value=f"{trade_red} fewer trades")
    row += 2

    # Key advantages
    ws.cell(row=row, column=1, value="Key DEVPRO SIP Advantages:")
    ws.cell(row=row, column=1).font = SECTION_FONT
    row += 1
    for adv in summary.get("key_advantages", []):
        ws.cell(row=row, column=1, value=f"• {adv}")
        row += 1

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def generate_excel(design, comparison=None, output_path="nzs3604_report.xlsx"):
    """Generate the Excel report workbook.

    Args:
        design: NZS 3604 engine output dict.
        comparison: Comparison report dict (optional).
        output_path: Output file path.
    """
    if not HAS_OPENPYXL:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()

    # Sheet 1: Timber Schedule
    ws1 = wb.active
    write_timber_schedule(ws1, design)

    # Sheet 2: Fixing Schedule
    ws2 = wb.create_sheet()
    write_fixing_schedule(ws2, design)

    # Sheet 3: Comparison (if available)
    if comparison:
        ws3 = wb.create_sheet()
        write_comparison(ws3, comparison)

    wb.save(output_path)
    print(f"Excel report saved to {output_path}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python nzs3604_excel.py nzs3604_design.json [comparison.json] [output.xlsx]")
        sys.exit(1)

    design_path = sys.argv[1]
    comparison_path = sys.argv[2] if len(sys.argv) > 2 and sys.argv[2].endswith(".json") else None
    output_path = sys.argv[-1] if sys.argv[-1].endswith(".xlsx") else "nzs3604_report.xlsx"

    with open(design_path, "r") as f:
        design = json.load(f)

    comparison = None
    if comparison_path:
        with open(comparison_path, "r") as f:
            comparison = json.load(f)

    generate_excel(design, comparison, output_path)


if __name__ == "__main__":
    main()
